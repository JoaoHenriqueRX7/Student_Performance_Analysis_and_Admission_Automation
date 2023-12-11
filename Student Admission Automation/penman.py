import json
import subprocess
import os
import zipfile
import pandas as pd
from faker import Faker
import datetime
from docxtpl import DocxTemplate
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Constants
MINIMUM_MATH_SCORE = 80
MINIMUM_READING_SCORE = 85
MINIMUM_WRITING_SCORE = 85
VACANCIES = 100
DATASET_NAME = 'rkiattisak/student-performance-in-mathematics'
RESOURCES_FOLDER = "project_assets"  
KAGGLE_CREDENTIAL_PATH = os.path.join(RESOURCES_FOLDER, "kaggle.json")
ADMISSION_TEMPLATE_PATH = os.path.join(RESOURCES_FOLDER, "admission_template.docx")

def read_kaggle_credentials():
    try:
        print("... loading credentials\n")
        with open(KAGGLE_CREDENTIAL_PATH, "r") as json_file:
            return json.load(json_file)
    except Exception as e:
        print(f"An Error has ocurred while reading the kaggle credentials: \n {e}")

def set_kaggle_credentials(credentials):
    os.environ['KAGGLE_USERNAME'] = credentials["username"]
    os.environ['KAGGLE_KEY'] = credentials["key"]

def download_dataset(download_path):
    try:
        print("... downloading dataset\n")
        subprocess.run(f'kaggle datasets download -d "{DATASET_NAME}" -p "{download_path}"', shell=True)
    except Exception as e:
        print(f"An Error has ocurred while downloading the dataset: \n {e}")

def extract_dataset(download_path, extraction_path):
    try:
        print("... extracting dataset\n")
        zip_file_path = os.path.join(download_path, 'student-performance-in-mathematics.zip')
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            zip_ref.extractall(extraction_path)
    except Exception as e:
        print(f"An Error has ocurred while extracting the dataset: \n {e}")

def load_student_data(extraction_path):
    csv_path = os.path.join(extraction_path, 'exams.csv')
    return pd.read_csv(csv_path)

def process_student_data(df):
    try:
        print("... processing Student data\n")
        df['total score'] = df['math score'] + df['reading score'] + df['writing score']
        df = df[df['math score'] >= MINIMUM_MATH_SCORE]
        df = df[df['reading score'] >= MINIMUM_READING_SCORE]
        df = df[df['writing score'] >= MINIMUM_WRITING_SCORE]
        return df.sort_values(by='total score', ascending=False).head(VACANCIES)
    except Exception as e:
        print(f"An Error has ocurred while extracting the dataset: \n {e}")

def add_fake_names(df):
    print("... adding fake names to preserve student's privacy\n")
    fake = Faker()
    df['name'] = [fake.name() for _ in range(len(df))]
    return df

def generate_admission_letters(admitted_students, output_admitted_folder):
    print("... generating letters to admitted students")
    date_now = datetime.datetime.now().strftime("%m/%d/%Y")
    doc = DocxTemplate(ADMISSION_TEMPLATE_PATH)
    letter_paths = []

    for _, row in admitted_students.iterrows():
        context = {"Student_Name": row['name'], "date": date_now}
        output_file = os.path.join(output_admitted_folder, f"{row['name']}_Welcome_Letter.docx")
        doc.render(context)
        doc.save(output_file)
        letter_paths.append(output_file)

    return letter_paths


def generate_admission_excel_list(admitted_students, letter_paths, output_folder):
    try:
        current_year = datetime.datetime.now().year
        filename = f"admitted_students_{current_year}.xlsx"
        output_path = os.path.join(output_folder, filename)

        # Add the letter paths to the DataFrame
        admitted_students['Welcome Letter'] = letter_paths

        # Save DataFrame to Excel
        admitted_students.to_excel(output_path, index=False)

        # Load the workbook and select the active worksheet
        workbook = openpyxl.load_workbook(output_path)
        worksheet = workbook.active

        # Apply styling and create hyperlinks
        for row, letter_path in enumerate(letter_paths, start=2):  # Start from row 2 (skip header)
            cell = worksheet.cell(row=row, column=len(admitted_students.columns))
            cell.hyperlink = letter_path
            cell.value = "Open Letter"
            cell.style = 'Hyperlink'

        # Set column widths based on content
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(output_path)
        print(f"Admitted students successfully saved to:\n {output_path}\n")
    except Exception as e:
        print(f"An error occurred while saving to Excel: \n{e}")
    

def main():
    kaggle_credentials = read_kaggle_credentials()
    set_kaggle_credentials(kaggle_credentials)

    project_folder = os.path.dirname(__file__)
    download_path = os.path.join(project_folder, 'download')
    extraction_path = os.path.join(project_folder, 'extracted_files')
    output_admitted_folder = os.path.join(project_folder, 'admitted_students')
    admission_lists_folder = os.path.join(project_folder, 'admission_lists')

    for path in [download_path, extraction_path, output_admitted_folder, admission_lists_folder]:
        if not os.path.exists(path):
            os.makedirs(path)

    download_dataset(download_path)
    extract_dataset(download_path, extraction_path)
    
    student_data = load_student_data(extraction_path)
    processed_data = process_student_data(student_data)
    admitted_students = add_fake_names(processed_data)
    letter_paths = generate_admission_letters(admitted_students, output_admitted_folder)
    
    
    generate_admission_excel_list(admitted_students, letter_paths, admission_lists_folder)
    
    print(f"Student letters successfully created at path: \n {output_admitted_folder}\n")

if __name__ == "__main__":
    main()
